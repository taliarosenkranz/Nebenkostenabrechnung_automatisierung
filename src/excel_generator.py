"""
═══════════════════════════════════════════════════════════════
EXCEL GENERATOR - Nebenkostenabrechnung
═══════════════════════════════════════════════════════════════
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import date
from typing import Dict, Any
import config


def create_nebenkostenabrechnung(
    data: Dict[str, Any],
    tenant_name: str,
    output_path: str,
    year: int,
    period_start: date = None,
    period_end: date = None
):
    """
    Erstellt Excel-Nebenkostenabrechnung
    
    Args:
        data: Berechnungsergebnis von calculate_tenant_costs()
        tenant_name: Name des Mieters
        output_path: Pfad zur Excel-Datei
        year: Abrechnungsjahr
        period_start: Startdatum des Abrechnungszeitraums (optional, default: 01.01.year)
        period_end: Enddatum des Abrechnungszeitraums (optional, default: 31.12.year)
    """
    # Set default period if not provided
    if period_start is None:
        period_start = date(year, 1, 1)
    if period_end is None:
        period_end = date(year, 12, 31)
    
    # Format period strings
    period_start_str = period_start.strftime('%d.%m.%Y')
    period_end_str = period_end.strftime('%d.%m.%Y')
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"Abrechnung {year}"
    
    # Column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 15
    
    # Styles
    header_font = Font(name='Arial', size=14, bold=True)
    subheader_font = Font(name='Arial', size=11, bold=True)
    normal_font = Font(name='Arial', size=10)
    bold_font = Font(name='Arial', size=10, bold=True)
    
    center_align = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    gray_fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
    blue_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    
    # Header
    row = 1
    
    # Landlord info
    ws[f'A{row}'] = config.LANDLORD['name']
    ws[f'A{row}'].font = normal_font
    row += 1
    
    ws[f'A{row}'] = config.LANDLORD['address']
    ws[f'A{row}'].font = normal_font
    row += 2
    
    # Tenant info
    ws[f'A{row}'] = f"Frau/Herr"
    ws[f'A{row}'].font = normal_font
    row += 1
    
    ws[f'A{row}'] = tenant_name
    ws[f'A{row}'].font = normal_font
    row += 1
    
    ws[f'A{row}'] = config.PROPERTY['address']
    ws[f'A{row}'].font = normal_font
    row += 3
    
    # Date
    ws[f'A{row}'] = f"Berlin, {date.today().strftime('%d.%m.%Y')}"
    ws[f'A{row}'].font = normal_font
    row += 2
    
    # Title
    ws[f'A{row}'] = f"Betriebskostenabrechnung {year}"
    ws[f'A{row}'].font = Font(name='Arial', size=16, bold=True)
    row += 1
    
    ws[f'A{row}'] = f"Heizkostenabrechnung {year}"
    ws[f'A{row}'].font = Font(name='Arial', size=16, bold=True)
    row += 2
    
    # Object info
    ws[f'A{row}'] = f"Objekt: Mieteinheit {config.PROPERTY['address']}"
    ws[f'A{row}'].font = normal_font
    row += 2
    
    # Intro text
    ws[f'A{row}'] = "Sehr geehrte/r Frau/Herr,"
    ws[f'A{row}'].font = normal_font
    row += 2
    
    ws[f'A{row}'] = (
        "gemäß § 3 des Mietvertrages sind die Heiz- und Warmwasserkosten sowie die Betriebskosten Ihrer"
    )
    ws[f'A{row}'].font = normal_font
    row += 1
    
    ws[f'A{row}'] = (
        f"Einheit in der Miete nicht enthalten, sondern werden separat abgerechnet. Ich erlaube mir daher, für"
    )
    ws[f'A{row}'].font = normal_font
    row += 1
    
    ws[f'A{row}'] = (
        f"den Zeitraum {period_start_str} - {period_end_str} die Heiz- und Warmwasserkosten sowie die Betriebskosten"
    )
    ws[f'A{row}'].font = normal_font
    row += 1
    
    ws[f'A{row}'] = "nachfolgend abzurechnen."
    ws[f'A{row}'].font = normal_font
    row += 3
    
    # Cost table header
    ws[f'A{row}'] = "Betriebs-/Heizkostenabrechnung"
    ws[f'A{row}'].font = subheader_font
    row += 1
    
    ws[f'A{row}'] = f"Abrechnungszeitraum WEG {period_start_str} - {period_end_str}"
    ws[f'A{row}'].font = normal_font
    row += 2
    
    # Table headers
    ws[f'A{row}'] = "Kostenart"
    ws[f'A{row}'].font = bold_font
    ws[f'A{row}'].fill = gray_fill
    ws[f'A{row}'].border = thin_border
    
    ws[f'B{row}'] = "Betrag"
    ws[f'B{row}'].font = bold_font
    ws[f'B{row}'].alignment = right_align
    ws[f'B{row}'].fill = gray_fill
    ws[f'B{row}'].border = thin_border
    row += 1
    
    # Cost items
    for item in data['items']:
        ws[f'A{row}'] = item['name']
        ws[f'A{row}'].font = normal_font
        ws[f'A{row}'].border = thin_border
        
        ws[f'B{row}'] = f"{item['tenant_share']:.2f} €"
        ws[f'B{row}'].font = normal_font
        ws[f'B{row}'].alignment = right_align
        ws[f'B{row}'].border = thin_border
        row += 1
    
    # Total costs
    ws[f'A{row}'] = "Gesamtkosten"
    ws[f'A{row}'].font = bold_font
    ws[f'A{row}'].fill = blue_fill
    ws[f'A{row}'].border = thin_border
    
    ws[f'B{row}'] = f"{data['total_costs']:.2f} €"
    ws[f'B{row}'].font = bold_font
    ws[f'B{row}'].alignment = right_align
    ws[f'B{row}'].fill = blue_fill
    ws[f'B{row}'].border = thin_border
    row += 2
    
    # Prepayments
    ws[f'A{row}'] = "abzgl. Ist-Vorauszahlungen"
    ws[f'A{row}'].font = normal_font
    row += 1
    
    ws[f'A{row}'] = f"{data['payment_months']} x Vorauszahlungen"
    ws[f'A{row}'].font = normal_font
    ws[f'A{row}'].border = thin_border
    
    ws[f'B{row}'] = f"-{data['prepayments']:.2f} €"
    ws[f'B{row}'].font = normal_font
    ws[f'B{row}'].alignment = right_align
    ws[f'B{row}'].border = thin_border
    row += 2
    
    # Balance
    balance = data['balance']
    balance_label = "Nachzahlung" if balance > 0 else "Guthaben"
    
    balance_fill = PatternFill(
        start_color='FFEB9C' if balance > 0 else 'C6EFCE',
        end_color='FFEB9C' if balance > 0 else 'C6EFCE',
        fill_type='solid'
    )
    
    ws[f'A{row}'] = balance_label
    ws[f'A{row}'].font = Font(name='Arial', size=12, bold=True)
    ws[f'A{row}'].fill = balance_fill
    ws[f'A{row}'].border = thin_border
    
    ws[f'B{row}'] = f"{abs(balance):.2f} €"
    ws[f'B{row}'].font = Font(name='Arial', size=12, bold=True)
    ws[f'B{row}'].alignment = right_align
    ws[f'B{row}'].fill = balance_fill
    ws[f'B{row}'].border = thin_border
    row += 3
    
    # Footer
    ws[f'A{row}'] = "Die Wohn-/Hausgeldabrechnung der Wohnungseigentümergemeinschaft ist in"
    ws[f'A{row}'].font = Font(name='Arial', size=9)
    row += 1
    
    ws[f'A{row}'] = "der Anlage in Kopie beigefügt. Die Belege dazu können nach vorheriger Termin-"
    ws[f'A{row}'].font = Font(name='Arial', size=9)
    row += 1
    
    ws[f'A{row}'] = "abstimmung bei der Hausverwaltung eingesehen werden."
    ws[f'A{row}'].font = Font(name='Arial', size=9)
    row += 3
    
    ws[f'A{row}'] = "Mit freundlichen Grüßen,"
    ws[f'A{row}'].font = normal_font
    row += 2
    
    ws[f'A{row}'] = config.LANDLORD['name']
    ws[f'A{row}'].font = normal_font
    
    # Save
    wb.save(output_path)
